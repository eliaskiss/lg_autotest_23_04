from openpyxl import load_workbook
from datetime import datetime, timedelta
from icecream import ic
import time

ic.configureOutput(includeContext=True)

file_name = 'trasfer_fee.xlsx'

# Open Excel File
wb = load_workbook(file_name,
                   read_only=True, # Read Only가 True이면, Lazy Loading이 발생해서 반드시 Close 해줘야만 핸드을 놓게됨
                   data_only=True) # Data Only가 True이면, cell의 함수 결과값을 가져옴

# Get Sheet list
ws_list = wb.sheetnames
# ic(ws_list)

# Select Worksheet
# ws = wb.active # 현재 활성화되어 있는 sheet(저장할때 선택된 시트)
ws = wb['항공요금']
# ws = wb[ws_list[0]]  # 첫번째 워크시트
# ws = wb[ws_list[-1]] # 마지막번째 워크시트

# Get Cell Value
cell_A1 = ws['A1']
# ic(cell_A1)
# ic(cell_A1.value)

# Get Formular Cell Value
cell_D2 = ws['D2']
# ic(cell_D2.value) # data_only의 값에 따라 다른 Cell값을 가져옴, data_only : True --> 8.43, data_only : False --> =A2+B2

# Get Datetime Cell Value
cell_F2 = ws['F2']
# ic(cell_F2.value)
cell_F2_value = cell_F2.value.strftime('%Y-%m-%d %H:%M:%S')
# ic(cell_F2_value)

# # Get All 1st Column
# index = 1
# while True:
#     cell = f'A{index}' # A1, A2, A3, ....
#     cell_value = ws[cell].value
#
#     if cell_value == '' or cell_value is None:
#         break
#
#     ic(cell, cell_value)
#     index += 1

# # Get All 2nd Row
# index = 0
# while True:
#     cell = f'{chr(ord("A") + index)}2' # ord("A") + 1: 42 + 1 = 43, chr(43) --> B : A2, B2, C2, ...
#     cell_value = ws[cell].value
#
#     if cell_value == '' or cell_value is None:
#         break
#
#     ic(cell, cell_value)
#     index += 1

# # Get row data
# index = 0
# for row in ws.rows:
#     if index != 0:
#         ic(index, f'Weigh:{row[0].value}', f'Fee:{row[1].value}')
#     index += 1

# Get row data 2nd way
index = 1
for row in ws.iter_rows(min_row=2):
    ic(index, f'Weigh:{row[0].value}', f'Fee:{row[1].value}')
    index += 1

# read_only=False 인 경우
wb.close()















