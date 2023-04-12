from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
from openpyxl.chart import BarChart, Reference

# Create New Excel object
wb = Workbook()

file_name = './new_excel.xlsx'

# Get Active Worksheet
ws = wb.active # 'sheet1'

#######################################################################################################################
# Set Sheet Name
ws.title = 'Basic'

# Set Cell Value
ws['A1'] = 'Hello World'
ws['B1'] = 10
ws['C1'] = 20

# Insert Comment
comment = Comment('This is comment', 'Elias Kim', 100, 100)
ws['A1'].comment = comment

# Insert Row
row = [1,2,3,4,5]
ws.append(row)
# for i in range(10):
#     ws.append(row)

# Use Function
ws['D1'] = '=SUM(B1+C1)'

# Insert Datetime
ws['A4'] = datetime.now()
ws['B4'] = datetime.now() + timedelta(days=1) # 내일
# ws['B4'] = datetime.now() - timedelta(days=1) # 어제
# ws['B4'] = datetime.now() + timedelta(hours=1) # 한시간 후
ws['C4'] = '2023-04-12'

# Merge Target Cell
ws['F1'] = 'Hello'
ws['G1'] = 'World'

# # Merge/Unmerge
# ws.merge_cells(range_string='F1:G1')
# ws.unmerge_cells('F1:G1')

# Merge/Unmerge II
ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
ws.unmerge_cells(start_row=1, start_column=6, end_row=1, end_column=7)

# Insert Image
img = Image('buz.jpg')
ws.add_image(img, 'B10')

# Create new sheet
wb.create_sheet('Insert Delete Move')

#######################################################################################################################
# Select new sheet
ws = wb['Insert Delete Move']

# Creat Row Data
row  = [x for x in range(20)] # [0,1,2,3,...,19]

for _ in range(20):
    ws.append(row)

# # Insert Empty Row
# ws.insert_rows(2)
#
# # Insert Empty Column
# ws.insert_cols(2)
#
# # Delete Rows
# ws.delete_rows(5, 7) # From 5, Step 7
#
# # Delete Columns
# ws.delete_cols(5, 7) # From 5, Step 7

# Move Row
ws.move_range('A1:U1', rows=15, cols=2) # Step 15 rows, Step 2 cols

#######################################################################################################################
# Doc Site: https://openpyxl.readthedocs.io/en/stable/charts/introduction.html

# Create & Seelct chart sheet
wb.create_sheet('Chart')
ws = wb['Chart']

for i in range(10):
    ws.append([i])

values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, 'C1')

#######################################################################################################################

from openpyxl.styles import Font, Color, PatternFill, GradientFill, Alignment, Side, Border

# Create New Sheet
wb.create_sheet('Style')
ws = wb['Style']

# Set Color
red_font = Font(color='FF0000')
ws['A1'] = 'Red'
ws['A1'].font = red_font

# Set Combination
combi_font = Font(color='395B64', size=20, bold=True, italic=True,
                  underline='singleAccounting', # underline: single, double, singleAccounting, doubleAccounting
                  strike=True)
ws['A2'] = 'Combination'
ws['A2'].font = combi_font

# Font Name
arial_font = Font(name='arial', size=14)
ws['A3'] = 'Arial'
ws['A3'].font = arial_font

# Cell Border
thin = Side(border_style='thin', color='000000')
ws['B4'] = 'Thin'
ws['B4'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

double = Side(border_style='double', color='000000')
ws['B6'] = 'Double'
ws['B6'].border = Border(top=double, left=double, right=double, bottom=double)


















wb.save(filename=file_name)
