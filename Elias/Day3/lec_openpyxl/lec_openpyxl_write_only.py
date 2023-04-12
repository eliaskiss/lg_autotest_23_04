# pip install pillow
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime

#######################################################################################
# Create Excel Object
wb = Workbook(
    write_only=True
)

file_name = './write_only_excel.xlsx'

# Get Ative Worksheet
ws = wb.create_sheet('My Data')

# Insert Row
row = [1,2,3,4,5]
ws.append(row)

# Insert Image
img = Image('./buz.jpg')
ws.add_image(img, 'A10')

# Create Sheet
wb.create_sheet('Extra')

# Select New Sheet
ws = wb['Extra']
# ws = wb[wb.sheetnames[1]] # or

ws.append([1,2,3,4,5])

from openpyxl.styles import Font
# Cell에 옵션을 줘서 쓰려면 에러
# Error!!!
# red_font = Font(color='FF0000')
# ws['A1'] = 'Red'
# ws['A1'].font = red_font

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
cell = WriteOnlyCell(ws, value="Red")
cell.font = Font(color='FF0000')
cell2 = WriteOnlyCell(ws, value="Green")
cell2.font = Font(color='00FF00')
ws.append([cell, cell2])

wb.save(filename=file_name)